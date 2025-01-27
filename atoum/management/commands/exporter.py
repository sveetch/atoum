"""
Command to export Atoum data into a XSLX file.
"""
import csv
import datetime
import json
import tempfile
from pathlib import Path

from django.core.management.base import BaseCommand
from django.utils.text import slugify

import tablib
from import_export import resources
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.packaging.core import DocumentProperties

from atoum.admin import (
    AssortmentResource, CategoryResource, ConsumableResource, ProductResource
)
from atoum.models import Assortment, Category, Consumable, Product


class Command(BaseCommand):
    """
    Export relevant Atoum data into a XSLX file where each data type (Consumable,
    Assortment, etc..) is in its own sheet.

    Attributes:
        EXPORT_MODELS (list): List of tuple for all model to export with import-export
            resource class. The list order will define the order of created sheets in
            the XSLX document.
    """
    # Models are listed in order of priority for the XLSX sheet to create
    EXPORT_MODELS = [
        (Consumable, ConsumableResource),
        (Assortment, AssortmentResource),
        (Category, CategoryResource),
        # ("Brand", BrandResource),
        (Product, ProductResource),
    ]

    def add_arguments(self, parser):
        pass

    def autofit(self, sheet):
        """
        Iterate each cell of each column to find the maximum length per column and
        adjust cells to fit to the maximum column length.
        """
        for column_cells in sheet.columns:
            max_column_length = max([len(cell.value or "") for cell in column_cells])

            # Adjust length for an extra space but finally limit to 250 units
            max_column_length = max_column_length + 2
            max_column_length = 250 if max_column_length > 250 else max_column_length

            # set the width of the column to the max_column_length
            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = max_column_length

    def format_header_cell(self, cell):
        """
        Define "header like" styles on given cell.

        This is expected to be use on the first row of a sheet.
        """
        # Shared colors
        light_blue = "0099CCFF"
        black = "00000000"
        white = "00FFFFFF"

        # Main border style
        thin_border = Side(border_style="thin", color=black)

        # Apply a background fill
        cell.fill = PatternFill(
            start_color=light_blue,
            end_color=light_blue,
            fill_type="solid"
        )

        # Apply the font styles
        cell.font = Font(name="Tahoma", size=12, color=black, bold=True)

        # Apply an alignment
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # Apply the border
        cell.border = Border(
            top=thin_border,
            left=thin_border,
            right=thin_border,
            bottom=thin_border
        )


    def export_model(self, item, csvdir):
        """
        Export model data into a CSV file.
        """
        model = item[0]
        resource = item[1]
        name = model.__name__
        filename = "{name}_{date}.csv".format(
            name=name,
            date=self.now.isoformat().split("T")[0],
        )
        destination = csvdir / filename

        self.stdout.write(
            "- {name} object(s) to export: {count}".format(
                name=name,
                count=model.objects.all().count(),
            )
        )
        dataset = resource().export()

        destination.write_text(dataset.csv)

        return destination

    def create_model_sheet(self, index, model, path):
        """
        Create a sheet in workbook with exported model data from CSV file.
        """
        model_name = str(model._meta.verbose_name_plural)
        # For first model, rename the current active sheet automatically created from
        # openpyxl
        if index == 0:
            sheet = self.workbook.active
            sheet.title = model_name
        # Else create a new sheet
        else:
            sheet = self.workbook.create_sheet(model_name)

        self.stdout.write("  └─ Written into sheet '{}'".format(model_name))

        with path.open() as f:
            reader = csv.reader(f, delimiter=",")

            for row_index, row in enumerate(reader, start=1):
                for column_index, cell_value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=column_index).value = cell_value

                self.autofit(sheet)

                for cell in sheet[1]:
                    self.format_header_cell(cell)

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS("=== XSLX Export ==="))
        self.now = datetime.datetime.now()
        self.workbook = Workbook()

        with tempfile.TemporaryDirectory() as tmpdirname:
            tempdir = Path(tmpdirname)
            for i, item in enumerate(self.EXPORT_MODELS):
                destination = self.export_model(item, tempdir)
                self.stdout.write("  └─ Written to temporary file '{}'".format(destination))

                self.create_model_sheet(i, item[0], destination)
        xslx_filepath = "Atoum_{}.xlsx".format(self.now.isoformat().split("T")[0])
        self.workbook.save(xslx_filepath)

        self.stdout.write("- Written workbook file '{}'".format(xslx_filepath))
